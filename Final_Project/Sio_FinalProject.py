import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl as op

main = tk.Tk()
main.title("Printing Service Ordering System")
main.geometry("500x300")
main.configure(bg="white")
main.resizable(False, False)

prices = {"Document Printing": 4.0, "Photocopying": 2.0, "Scanning": 5.0, "Lamination": 15.0, "Binding": 50.0, "Photo Printing": 20.0, "ID Picture": 40.0}
size_price = {"A4": 1.0, "A3": 1.5, "A5": 0.8, "Letter": 1.2, "Legal": 1.3, "Photo": 2.0}
color_price = {"Black & White": 1.0, "Color": 1.5, "Full Color": 2.0}

#maam pa change nlng po dun sa loc ng DataBase.xlsx sainyo 
file = '/home/sam/Downloads/Sio_Code/Final_Project/Sio_DataBase.xlsx'

def calculate_price(event=None):
    try:
        service = doc_drop.get()
        size = size_drop.get()
        color = color_drop.get()

        pages = int(page_entry.get())
        copies = int(copy_entry.get())

        base_price = prices.get(service, 0)
        size_factor = size_price.get(size, 1)
        color_factor = color_price.get(color, 1)

        total = pages * copies * base_price * size_factor * color_factor

        price_display.config(text=f"₱{total:.2f}")
        return total

    except:
        price_display.config(text="₱0.00")
        return 0

def show_treeview():
    global table

    excel = tk.Toplevel()
    excel.title("InkHub Service Orders")
    excel.geometry("850x400")
    excel.configure(bg="white")
    excel.resizable(False, False)

    # Sidebar
    frame = tk.Frame(excel, bg="gray", width=50, height=400)
    frame.place(x=0, y=0)

    # Title
    title = tk.Label(excel, text="InkHub Service Orders", font=("times new roman", 15, "bold"), bg="white")
    title.place(x=60, y=10)

    # Subtitle
    subtitle = tk.Label(excel, text="View and manage customer orders", font=("times new roman", 10), bg="white")
    subtitle.place(x=60, y=40)

    # Separator line
    border = tk.Frame(excel, bg="light gray", height=1, width=720)
    border.place(x=60, y=65)

    # Treeview
    columns = ("Order ID", "Client Name", "Service", "Paper Size", "Paper Color", "Pages", "Copies", "Price")

    table = ttk.Treeview(excel, columns=columns, show="headings")

    for heading in columns:
        table.heading(heading, text=heading)

    table.column("Order ID", width=70)
    table.column("Client Name", width=140)
    table.column("Service", width=140)
    table.column("Paper Size", width=80)
    table.column("Paper Color", width=100)
    table.column("Pages", width=60)
    table.column("Copies", width=60)
    table.column("Price", width=80)

    # Scrollbar
    scrollbar = ttk.Scrollbar(excel, orient="vertical", command=table.yview)
    table.configure(yscrollcommand=scrollbar.set)
    table.place(x=60, y=80, width=750, height=250)
    scrollbar.place(x=810, y=80, height=250)

    # Footer
    total_label = tk.Label(excel, text="Total Orders: 0", font=("times new roman", 10, "bold"), bg="white")
    total_label.place(x=60, y=350)

    # Close button
    close_btn = tk.Button(excel, text="Close", font=("times new roman", 10), width=10, command=excel.destroy)
    close_btn.place(x=680, y=345)

    table.bind("<<TreeviewSelect>>", auto_populates)
    wordings()

def wordings():
    workbook = op.load_workbook(file)
    order = workbook.active

    for content in table.get_children():
        table.delete(content)

    for rows in order.iter_rows(min_row=2,values_only= True):
        table.insert("",tk.END,values= rows)

def auto_populates(event):
    selected = table.focus()
    values = table.item(selected, "values")
    
    if values:
        name_entry.delete(0,tk.END)
        page_entry.delete(0,tk.END)
        copy_entry.delete(0,tk.END)

        name_entry.insert(0,values[1])
        doc_drop.set(values[2])
        size_drop.set(values[3])
        color_drop.set(values[4])
        page_entry.insert(0,values[5])
        copy_entry.insert(0,values[6])

def validation():
    client = name_entry.get()
    service = doc_drop.get()
    size = size_drop.get()
    color = color_drop.get()
    pages = page_entry.get()
    copies = copy_entry.get()

    #Required fields
    if service == "Select Service":
        messagebox.showerror("Error", "Please select a service")
        return False

    if size == "Select Type":
        messagebox.showerror("Error", "Please select a paper size")
        return False

    if color == "Select Type":
        messagebox.showerror("Error", "Please select a paper color")
        return False
    
    if not client or not pages or not copies:
        messagebox.showerror("Error", "All fields are required!")
        return False
    
    #Must be numeric
    if not pages.isdigit() or not copies.isdigit():
        messagebox.showerror("Error","Pages and Copies shoould be a number!")
        return False
    
    return True

def get_next_id(sheet):
    ids = [
        int(r[0])
        for r in sheet.iter_rows(min_row=2, values_only=True)
        if r[0] is not None and str(r[0]).isdigit()
    ]
    return max(ids, default=0) + 1

def Order():
    if not validation():
        return
    
    client = name_entry.get()
    service = doc_drop.get()
    size = size_drop.get()
    color = color_drop.get()
    pages = int(page_entry.get())
    copies = int(copy_entry.get())
    price = calculate_price()
    
    workbook = op.load_workbook(file)
    order = workbook.active

    order_id = get_next_id(order)

    order.append([order_id, client, service, size, color, pages, copies, price])
    workbook.save(file)
    
    messagebox.showinfo("Success","Record added successfully!")
    wordings()

def update():
    selected = table.focus()

    if not selected:
        messagebox.showerror("Error", "Select a record first")
        return

    if not validation():
        return
    
    values = table.item(selected, "values")
    order_id = int(values[0])

    client = name_entry.get()
    service = doc_drop.get()
    size = size_drop.get()
    color = color_drop.get()
    pages = int(page_entry.get())
    copies = int(copy_entry.get())
    price = calculate_price()
    

    workbook = op.load_workbook(file)
    order = workbook.active

    for row in order.iter_rows(min_row=2):
        if row[0].value == order_id:
            row[1].value = client
            row[2].value = service
            row[3].value = size
            row[4].value = color
            row[5].value = pages
            row[6].value = copies
            row[7].value = price

    workbook.save(file)

    messagebox.showinfo("Success", "Record updated Successfully!")
    wordings()

def delete():
    selected = table.focus()

    if not selected:
        messagebox.showerror("Error", "Select a record first")
        return
    
    values = table.item(selected, "values")
    order_id = int(values[0])

    confirm = messagebox.askyesno("Confirm", "Are you sure you want to delete this record?")
    if not confirm:
        return 
    
    workbook = op.load_workbook(file)
    order = workbook.active

    for i, row in enumerate(order.iter_rows(min_row=2), start=2):
        if row[0].value == order_id:
            order.delete_rows(i)
            break
    workbook.save(file)

    messagebox.showinfo("Success", "Record deleted Successfully!")
    wordings()

#black design sa tabi
frame = tk.Frame(main, bg= "LightSteelBlue4", width= 50, height= 300)
frame.place(x= 0, y= 0)

#title
title = tk.Label(main, text="Welcome to InkHub Service!", font= ("times new roman", 15, "bold"), bg="white")
title.place(x= 55, y= 10)

#line
border = tk.Frame(main, bg="light gray", height=1, width=430)
border.place(x=55, y=65)

#instruction for picking
ins_label = tk.Label(main, text= "Enter your name and select the services you wish to order.", font= ("times new roman", 10), bg= "white")
ins_label.place(x= 55, y= 40)

#name of client
name_label = tk.Label(main, text= "Name of Client", font= ("times new roman", 10, "bold"), bg= "white")
name_label.place(x= 55, y= 75)

name_entry = tk.Entry(main, font= ("times new roman", 10), bg= "white", width= 20, relief= "flat")
name_entry.place(x= 55, y= 95)

#paper size 
size_label = tk.Label(main, text= "Paper Size", font= ("times new roman", 10, "bold"), bg= "white")
size_label.place(x= 200, y= 135)

size_drop = ttk.Combobox(main, values= ["A4", "A3", "A5", "Letter", "Legal", "Photo"], font= ("times new roman", 10), background= "white", width= 15, state= "readonly")
size_drop.set("Select Type")
size_drop.place(x= 200, y= 155)

#paper color
color_label = tk.Label(main, text= "Paper Color", font= ("times new roman", 10, "bold"), bg= "white")
color_label.place(x= 345, y= 135)

color_drop = ttk.Combobox(main, values= ["Black & White", "Full Color", "Color"], font= ("times new roman", 10), background= "white", width= 15, state= "readonly")
color_drop.set("Select Type")
color_drop.place(x= 345, y= 155)

#document title
doc_label = tk.Label(main, text= "Service", font= ("times new roman", 10, "bold"), bg= "white")
doc_label.place(x= 55, y= 135)

#drop box for document
doc_drop = ttk.Combobox(main, values= ["Document Printing", "Photocopying", "Scanning", "Lamination", "Binding", "Photo Printing", "ID Picture"], font= ("times new roman", 10), background= "white", width= 15, state= "readonly")
doc_drop.set('Select Service') 
doc_drop.place(x=55, y= 155)

#Entry for Pages
page_label = tk.Label(main, text= "Pages", font= ("times new roman", 10, "bold"), bg= "white")
page_label.place(x= 55, y= 190)

page_entry = tk.Entry(main, font= ("times new roman", 10), bg= "white", width= 5, relief= "flat")
page_entry.place(x= 55, y= 210)

#Entry for Copy
copy_label = tk.Label(main, text= "Copies", font= ("times new roman", 10, "bold"), bg= "white")
copy_label.place(x= 120, y= 190)

copy_entry = tk.Entry(main, font= ("times new roman", 10), bg= "white", width= 5, relief= "flat")
copy_entry.place(x= 120, y= 210)

# Price
price_label = tk.Label(main, text="Price", font=("times new roman", 11, "bold"), bg="white")
price_label.place(x=185, y=200)

price_display = tk.Label(main, text="₱0.00", font=("times new roman", 11), bg="white")
price_display.place(x= 250, y=200)

#Button for Edit
edit_btn = tk.Button(main, text= "Update", font= ("times new roman", 10), bg= "light gray", relief= "raised", width= 10, height= 1,command= update)
edit_btn.place(x= 55, y= 250)

#Button for Delete
del_btn = tk.Button(main, text= "Delete", font= ("times new roman", 10), bg= "light gray", relief= "raised", width= 10, height= 1, command= delete)
del_btn.place(x= 165, y= 250)

#Button to show Treeview
tree_btn = tk.Button(main, text= "View Orders",  font= ("times new roman", 10), bg= "light gray", relief= "raised",width= 10,  height= 1, command= show_treeview)
tree_btn.place(x= 275, y= 250)

#Button for Order
order_btn = tk.Button(main, text= "Add Order", font= ("times new roman", 10), bg= "light gray", relief= "raised", width= 10, height= 1, command= Order)
order_btn.place(x= 385, y= 250)

page_entry.bind("<KeyRelease>", calculate_price)
copy_entry.bind("<KeyRelease>", calculate_price)
doc_drop.bind("<<ComboboxSelected>>", calculate_price)
size_drop.bind("<<ComboboxSelected>>", calculate_price)
color_drop.bind("<<ComboboxSelected>>", calculate_price)

main.mainloop()
