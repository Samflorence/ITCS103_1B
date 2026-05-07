###Creating a workbook 
import openpyxl as op

wbk = op.Workbook
sheet = wbk.active

#writing on a workbook
sheet["A1"] = "Product"
sheet["B1"] = "Price"
sheet["C1"] = "Quantity"

sheet["A2"] = "Pensil"
sheet["B2"] = 15
sheet["C2"] = 3

wbk.save("try.xlsx")

###reading a workbook
import openpyxl as op

wbk = op.load_workbook('try.xlsx')
sheet = wbk.active

#read all collumns iter_cols
for columns in sheet.iter_cols(values_only = True):
    print(columns)

#readm all rows iter_rows
for columns in sheet.rows(values_only = True):
    print(rows)

#read specific cell
price = sheet["B2"].value
print (price)