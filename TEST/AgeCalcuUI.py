import tkinter as tk
import openpyxl as op
from tkinter import messagebox,ttk

def wordings():
    book = op.load_workbook("excelDb.xlsx")
    sheet = book.active

    #clear table
    for content in tree.get_children():
        tree.delete(content)

    for rows in sheet.iter_rows(min_row=2,values_only = True):
        tree.insert("",tk.END,value=rows)

def validation():
    first = fname_entry.get()
    last = lname_entry.get()
    by = birth_entry.get()

    #Required fields
    if not first or not last or not by:
        messagebox.showerror("Error", "All fields are required!")
        return False
    
    #Must be numeric
    if not by.isdigit():
        messagebox.showerror("Error","Birthyear shoould be a number!")
        return False
    
    return True

def appeand_excel():
    if not validation():
        return
    
    first = fname_entry.get()
    middle = mname_entry.get()
    last = lname_entry.get()
    by = int(birth_entry.get())
    age = 2026 - by 
    
    book = op.load_workbook("excelDb.xlsx")
    sheet = book.active

    new_id = sheet.max_row

    sheet.append([new_id, last, first, middle, by, age])
    book.save("excelDb.xlsx")
    
    messagebox.showinfo("Success","Record added successfully!")
    wordings()

def auto_populate():
    selected = tree.focus()
    values = tree.item(selected, "values")
    
    if values:
        lname_entry.delete(0,tk.END)
        fname_entry.delete(0,tk.END)
        mname_entry.delete(0,tk.END)
        birth_entry.delete(0,tk.END)

        lname_entry.insert(0,values[1])
        fname_entry.insert(0,values[2])
        mname_entry.insert(0,values[3])
        birth_entry.insert(0,values[4])
        
def update_data():
    selected = tree.focus()

    if not selected:
        messagebox.showerror("Error", "Select a record first")
        return

    if not validation():
        return
    
    values = tree.item(selected, "values")
    record_id = int(values[0])

    first = fname_entry.get()
    middle = mname_entry.get()
    last = lname_entry.get()
    by = int(birth_entry.get())
    age = 2026 - by

    workbook = op.load_workbook("excelDb.xlsx")
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2):
        if row[0].value == record_id:
            row[1].value == last
            row[2].value == first
            row[3].value == middle
            row[4].value == by
            row[5].value == age

    workbook.save("excelDb.xlsx")

    messagebox.showinfo("Success", "Record updated Successfully!")
    wordings()

def delete_data():
    selected = tree.focus()

    if not selected:
        messagebox.showerror("Error", "Select a record first")
        return
    
    values = tree.item(selected, "values")
    record_id = int(values[0])

    confirm = messagebox.askyesno("Confirm", "Are you sure you want to delete this record?")
    if not confirm:
        return 
    
    workbook = op.load_workbook("excelDb.xlsx")
    sheet = workbook.active

    for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if row[0].value == record_id:
            sheet.delete_rows(i)
            break
    workbook.save("excelDb.xlsx")

    messagebox.showinfo("Success", "Record deleted Successfully!")
    wordings()

window=tk.Tk()
window.title("Age Calculator")
window.configure(bg="lightgreen")


#Form Title
title = tk.Label ( window, text="Profile Builder", font=("Times New Roman",14,"bold"),bg="lightgreen")
title.grid(row=0, column=0, columnspan=6)

#Frame
genframe = tk.Frame(window,bg="lightgreen",bd=2, relief="groove")
genframe.grid(row=1,column=0, columnspan=6,padx=10,pady=10)

#First Name Entry
fname_entry = tk.Entry(genframe, font=("Poppins",12))
fname_entry.grid(row=2, column=1,columnspan=2,padx=(10,0),pady=(10,0))

fname_label = tk.Label(genframe, text="First Name", font=("Poppins",10,"italic"),bg="lightgreen")
fname_label.grid(row=3, column=1,columnspan=2)

#Middle Name Entry
mname_entry = tk.Entry(genframe, font=("Poppins",12))
mname_entry.grid(row=2, column=3,columnspan=2,padx=(10,0),pady=(10,0))

mname_label = tk.Label(genframe, text="Middle Name", font=("Poppins",10,"italic"),bg="lightgreen")
mname_label.grid(row=3, column=3,columnspan=2)

#Last Name Entry
lname_entry = tk.Entry(genframe, font=("Poppins",12))
lname_entry.grid(row=2, column=5,columnspan=2,padx=(10,10),pady=(10,0))

lname_label = tk.Label(genframe, text="Last Name", font=("Poppins",10,"italic"),bg="lightgreen")
lname_label.grid(row=3, column=5,columnspan=2)

#Birthyear Entry
birth_entry = tk.Entry(genframe, font=("Poppins",12))
birth_entry.grid(row=4, column=1,columnspan=2,padx=(10,0))

birthyear_label = tk.Label(genframe, text="Birth Year", font=("Poppins",10,"italic"),bg="lightgreen")
birthyear_label.grid(row=5, column=1,columnspan=2)

update_btn = tk.Button(window, text="Update",command=update_data)
update_btn.grid(row=6, column=2)

button= tk.Button(window,text="Submit",font=("Poppins",12,"bold"),bg="lightpink",command=appeand_excel)
button.grid(row=6, column=0, columnspan=6,pady=(10,20))

delete_btn = tk.Button(window, text="Delete",  bg="red", fg="white",command= delete_data)
delete_btn.grid(row=6, column=3)

tree = ttk.Treeview(window, columns=("ID","Last","First","Middle","BirthYear","Age"), show="headings")
for col in ("ID","Last","First","Middle","BirthYear","Age"):
    tree.heading(col, text=col)
tree.grid(row=7, column=0, columnspan=4)

tree.bind("<<Treeviewselect>>", auto_populate)
wordings()

window.mainloop()
